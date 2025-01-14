from itertools import takewhile
from operator import itemgetter

import click
import openpyxl

from hexagonal.utils import nettoyer_avec_spec

NOMS_COLONNES = {
    "PMUN": "population_municipale",
    "PSDC": "population_sans_doubles_comptes",
    "PTOT": "population_totale",
}


def entier(i):
    if i is None:
        return
    try:
        return int(i)
    except TypeError:
        return int(round(float(i), 0))


@click.command()
@click.argument("source", type=click.Path(exists=True, dir_okay=False, readable=True))
@click.argument("dest", type=click.Path(dir_okay=False, writable=True))
def run(source, dest):
    workbook = openpyxl.load_workbook(source, read_only=True, data_only=True)
    data_worksheet = workbook.worksheets[0]
    data_worksheet.reset_dimensions()  # l'INSEE a mal enregistr√© les dimensions
    columns = []

    i = 1
    while col := data_worksheet.cell(6, i).value:
        columns.append(col)
        i += 1

    colonnes_pop = [c for c in columns if c[:4] in ("PMUN", "PSDC", "PTOT")]

    pop_municipale = [c[-4:] for c in columns if c.startswith("PMUN")]
    pop_sans_double_compte = [c[-4:] for c in columns if c.startswith("PSDC")]
    pop_totale = [c[-4:] for c in columns if c.startswith("PTOT")]

    spec = {"code_commune": "CODGEO"}

    for col in colonnes_pop:
        type_pop = col[:4]
        annee = col[4:]

        nom_colonne = f"{NOMS_COLONNES[type_pop]}_{annee}"

        spec[nom_colonne] = (col, entier)

    rows = (
        {c: r[i].value for i, c in enumerate(columns)}
        for r in data_worksheet.iter_rows(min_row=7, max_col=len(columns))
    )
    rows = takewhile(itemgetter("CODGEO"), rows)

    nettoyer_avec_spec(rows, dest, spec)


if __name__ == "__main__":
    run()
